import os
from os.path import isfile, join
import sys

from openpyxl import Workbook
from openpyxl import load_workbook

from VBaseFunctions import *

#bPath="/Users/Stella/Documents/Ph.D. Research/Tool/Tool_Output_SWJ"
#bPath="/Users/Stella/Documents/Ph.D. Research/Tool/Tool_Output_LOV"
#bPath="/Users/Stella/Documents/Ph.D. Research/Tool/Single_Test_Output"
#bPath="/Users/Stella/Documents/Ph.D. Research/Base_URI_Tool/Base_URI_Tool_Output(SWJ)"
bPath="/Users/Stella/Documents/Ph.D. Research/Base_URI_Tool/Base_URI_Tool_Output(LOV)"
#bPath="YOUR OUTPUT FILE LOCATION HERE"
bFileName="vocab_Base.xlsx"
bFile=bPath+"/"+bFileName
bwb=load_workbook(filename=bFile)
bws=bwb["Vocab Data"]
mws=bwb["5-Star Vocabs"]
fws=bwb["From Links"]
maxRow=bws.max_row
nextVocabRow=2
while (nextVocabRow <= maxRow):
	getIndex=setIndices(nextVocabRow)
	pVocab=bws[getIndex.aInd].value
	pVocabStartRow=nextVocabRow
	fromLinksList=[]
	fBase5Star=False
	fProcessed=False
	pBVocabInfo=collectBVocabInfo(pVocab, bFile, bwb, bws, nextVocabRow)
	pVocabEndRow=pBVocabInfo.retRow-1
	nextVocabRow=pBVocabInfo.retRow
	currentRow=2

	while (currentRow < maxRow):
		if (currentRow==pVocabStartRow):
			currentRow=pVocabEndRow+1

		getIndex=setIndices(currentRow)
		currentBVocab=bws[getIndex.aInd].value
		if (bws[getIndex.dInd].value != None):
			if (bws[getIndex.dInd].value == "Processed"):
				fProcessed=True
		
		currentBVocabInfo=collectBVocabInfo(currentBVocab, bFile, bwb, bws, currentRow)
		fPresent=False
		for vocab in currentBVocabInfo.retLinksDict.keys():
			if (pVocab == vocab or (len(vocab) > len(pVocab) and vocab.find(pVocab) != -1)):
				fBase5Star=True
				uIndex='C'+str(pBVocabInfo.retBaseDict[pVocab])
				if (bws[uIndex] != "Yes"):
					bws[uIndex]="Yes"
				for link in fromLinksList:
					if (link == currentBVocab):
						fPresent=True
					else:
						fPresent=False
				if (not fPresent):
					fromLinksList.append(currentBVocab)

			for link in pBVocabInfo.retLinksDict.keys():
				processRow=2
				while (processRow <= mws.max_row):
					chkIndex='A'+str(processRow)
		
					if (link == mws[chkIndex].value):
						fBase5Star=True
						uIndex='C'+str(pBVocabInfo.retLinksDict[link])
						if (bws[uIndex] != "Yes"):
							bws[uIndex]="Yes"

						uIndex='C'+str(pBVocabInfo.retBaseDict[pVocab])
						if (bws[uIndex] != "Yes"):
							bws[uIndex]="Yes"	
						
					processRow += 1

			# 	for vocab in currentBVocabInfo.retLinksDict.keys():
			# 		if (link == vocab):
			# 			fBase5Star=True
			# 			uIndex='C'+str(pBVocabInfo.retBaseDict[pVocab])
			# 			if (bws[uIndex] != "Yes"):
			# 				bws[uIndex]="Yes"

			# 			uIndex='C'+str(pBVocabInfo.retLinksDict[link])
			# 			if (bws[uIndex] != "Yes"):
			# 				bws[uIndex]="Yes"

			# 			if (not fProcessed):
			# 				uIndex='C'+str(currentBVocabInfo.retBaseDict[currentBVocab])
			# 				if (bws[uIndex] != "Yes"):
			# 					bws[uIndex]="Yes"

			# 				uIndex='C'+str(currentBVocabInfo.retLinksDict[vocab])
			# 				if (bws[uIndex] != "Yes"):
			# 					bws[uIndex]="Yes"
		
		currentRow=currentBVocabInfo.retRow
		fProcessed=False

	if (fBase5Star):
		cFromLinkRow=fws.max_row + 1
		aInd='A'
		bInd='B'
		aIndex=aInd+str(cFromLinkRow)
		fws[aIndex]=pVocab

		if (fromLinksList):
			for item in fromLinksList:
				bIndex=bInd+str(cFromLinkRow)
				fws[bIndex]=item
				cFromLinkRow += 1
		else:
			bIndex=bInd+str(cFromLinkRow)
			fws[bIndex]="Is a 5-Star ontology/model because it reuses another 5-Star ontology/model"
			cFromLinkRow += 1

	dInd='D'
	dIndex=dInd+str(pVocabStartRow)
	bws[dIndex]="Processed"
bwb.save(bFile)