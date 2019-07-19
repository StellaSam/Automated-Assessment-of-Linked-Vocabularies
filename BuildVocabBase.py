from openpyxl import Workbook
from openpyxl import load_workbook

from openpyxl import Workbook
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.styles import Font

from CheckIfDuplicate import *

def buildBaseList(bFile):
	vBaseList=[]

	vwb=load_workbook(filename=bFile)
	vws=vwb.sheetnames	
	vws=vwb[vws[0]]
	cRow=2
	while (cRow <= vws.max_row):
		vIndex='A'+str(cRow)
		tIndex='B'+str(cRow)
		vInfo=[vws[vIndex].value, vws[tIndex].value]
		vBaseList.append(vInfo)
		cRow += 1
	
	return(vBaseList)

def addToVocabBase(bFile, vocab):
	vwb=load_workbook(filename=bFile)
	vws=vwb.sheetnames	
	vws=vwb[vws[0]]

	vInfo=[vocab, "Base"]
	vws.append(vInfo)
	vwb.save(bFile)

def createVocabBase(bFile, vocab):
	vwb=Workbook()
	vws=vwb.active
	vws.title="Vocab Data"
	vws1=vwb.create_sheet("5-Star Vocabs")
	vws2=vwb.create_sheet("From Links")

	vws.column_dimensions['A'].width=70
	vws.column_dimensions['B'].width=70
	vws.column_dimensions['C'].width=70
	vws.column_dimensions['D'].width=70

	vws['A1']="Vocabulary Name"
	vws['A1'].font=Font(bold=True)
	vws['B1']="Type"
	vws['B1'].font=Font(bold=True)
	vws['C1']="5-Star?"
	vws['C1'].font=Font(bold=True)
	vws['D1']="Processed?"
	vws['D1'].font=Font(bold=True)
	vws['A2']=vocab
	vws['B2']="Base"

	vws1.column_dimensions['A'].width=70
	vws1['A1']="Vocabulary Name"
	vws1['A1'].font=Font(bold=True)

	vws2.column_dimensions['A'].width=70
	vws2.column_dimensions['B'].width=70
	vws2['A1']="Vocabulary Name"
	vws2['A1'].font=Font(bold=True)
	vws2['B1']="From Links"
	vws2['B1'].font=Font(bold=True)

	vwb.save(bFile)
	
def addLinksToVocabBase(bFile, iBVocabTypeList, iDict):
	tempList=[]
	vwb=load_workbook(filename=bFile)
	vws=vwb.sheetnames	
	vws=vwb[vws[0]]

	for category, vocabList in iDict.items():
		for vocab in vocabList:
			if (not tempList):
				tempList.append(vocab)
				vInfo=[vocab, "Link"]
				iBVocabTypeList.append(vInfo)
			else:
				if (not checkIfDuplicate(vocab, tempList)):
					tempList.append(vocab)
					vInfo=[vocab, "Link"]
					iBVocabTypeList.append(vInfo)
					
	for vocab in (tempList):
		vInfo=[vocab, "Link"]
		vws.append(vInfo)
		
	vwb.save(bFile)
	
	return(iBVocabTypeList)
