from openpyxl import Workbook
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.styles import Font, Alignment

def writeIntoExcel(ws, writeValues, cRow):
	ws.title="4-Star Ranked Vocabs"
	ws.column_dimensions['A'].width=20
	ws.column_dimensions['B'].width=70
	ws.column_dimensions['C'].width=20
	ws.column_dimensions['D'].width=70
	ws.column_dimensions['E'].width=20
	ws.column_dimensions['F'].width=20

	ws['A1']="Vocabulary Name"
	ws['A1'].font=Font(bold=True)
	ws['A1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['B1']="Base URL"
	ws['B1'].font=Font(bold=True)
	ws['B1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['C1']="Links To Vocabulary"
	ws['C1'].font=Font(bold=True)
	ws['C1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['E1']="Has Meta-Data?"
	ws['E1'].font=Font(bold=True)
	ws['E1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['F1']="Star Rating"
	ws['F1'].font=Font(bold=True)
	ws['F1'].alignment=Alignment(horizontal='center', wrap_text='True')

	ws['A1']="Vocabulary Name"
	ws['B1']="Base URL"
	ws['C1']="Links To Vocabulary"
	ws['E1']="Has Meta-Data?"
	ws['F1']="Star Rating"

	vRow=cRow
	colRowA='A'+str(vRow)
	colRowB='B'+str(vRow)
	colRowC='C'+str(vRow)
	colRowD='D'+str(vRow)
	colRowE='E'+str(vRow)
	colRowF='F'+str(vRow)
	
	ws[colRowA]=writeValues.vocabFile
	ws[colRowA].alignment=Alignment(horizontal='center', wrap_text='True')
	ws[colRowB]=writeValues.baseOfVocab
	ws[colRowE]=writeValues.fhasMetaData
	ws[colRowE].alignment=Alignment(horizontal='center', wrap_text='True')
	ws[colRowF]=writeValues.starRate
	ws[colRowF].alignment=Alignment(horizontal='center', wrap_text='True')

	writeDict=writeValues.linksToDict
	row=vRow
	for key in writeDict:
		ws[colRowC]=key
		writeList=writeDict[key]
		for item in writeList:
			ws[colRowD]=item
			row += 1
			colRowD='D'+str(row)
		colRowC='C'+str(row)

	return(row)