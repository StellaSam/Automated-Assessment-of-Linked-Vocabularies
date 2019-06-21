from openpyxl.styles import Font, Alignment

from VBaseFunctions import *

def setIndices(row):
	colA='A'
	colB='B'
	colC='C'
	colD='D'
	colE='E'
	colF='F'

	class retIndices:
		def _init_ (self, aInd, bInd, cInd, dInd, eInd, fInd):
			self.aInd=""
			self.bInd=""
			self.cInd=""
			self.dInd=""
			self.eInd=""
			self.fInd=""

	retIndices.aInd=colA+str(row)
	retIndices.bInd=colB+str(row)
	retIndices.cInd=colC+str(row)
	retIndices.dInd=colD+str(row)
	retIndices.eInd=colE+str(row)
	retIndices.fInd=colF+str(row)

	return(retIndices)

def init5StarSheet(ws):
	ws.column_dimensions['A'].width=20
	ws.column_dimensions['B'].width=70
	ws.column_dimensions['C'].width=70
	ws.column_dimensions['D'].width=20
	ws.column_dimensions['E'].width=70
	ws.column_dimensions['F'].width=20

	ws['A1']="Vocabulary Name"
	ws['A1'].font=Font(bold=True)
	ws['A1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['B1']="Base URL"
	ws['B1'].font=Font(bold=True)
	ws['B1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['C1']="Links To Vocabulary"
	ws['C1'].font=Font(bold=True)
	ws['C1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['D1']="Has Meta-Data?"
	ws['D1'].font=Font(bold=True)
	ws['D1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['E1']="Links From Vocabulary"
	ws['E1'].font=Font(bold=True)
	ws['E1'].alignment=Alignment(horizontal='center', wrap_text='True')
	ws['F1']="Star Rating"
	ws['F1'].font=Font(bold=True)
	ws['F1'].alignment=Alignment(horizontal='center', wrap_text='True')

def write5StarSheet(ws, currentInfo, updateInfo, row):
	getIndex=setIndices(row)

	vToLinksList=currentInfo.retToLinksList
	vFromLinksList=updateInfo.retFromLinksList

	ws[getIndex.aInd]=currentInfo.retVName
	ws[getIndex.aInd].alignment=Alignment(horizontal='center', wrap_text='True')
	ws[getIndex.bInd]=currentInfo.retVocab
	ws[getIndex.dInd]=currentInfo.retFHasMeta
	ws[getIndex.dInd].alignment=Alignment(horizontal='center', wrap_text='True')
	ws[getIndex.fInd]=5
	ws[getIndex.fInd].alignment=Alignment(horizontal='center', wrap_text='True')

	currentRow=row
	for item in vToLinksList:
		ws[getIndex.cInd]=item
		currentRow += 1
		getIndex=setIndices(currentRow)
	toLinksRow=currentRow

	currentRow=row
	getIndex=setIndices(currentRow)
	for item in vFromLinksList:
		ws[getIndex.eInd]=item
		currentRow += 1
		getIndex=setIndices(currentRow)
	fromLinksRow=currentRow

	return(max(toLinksRow, fromLinksRow))

def collect4StarInfo(vocab, vFile, vwb, vws):
	maxRow=vws.max_row
	row=2
	getIndex=setIndices(row)

	class ret4StarInfo:
		def _init_ (self, retVName, retVocab, retToLinksList, retFHasMeta, retStarRate):
			self.retVName=""
			self.retVocab=""
			self.retToLinksList=[]
			self.retFHasMeta=False
			self.retStarRate=0

	while ((row <= maxRow) and (vocab != vws[getIndex.bInd].value)):
		row += 1
		getIndex=setIndices(row)

	if (vws[getIndex.bInd].value == vocab):
		vName=vws[getIndex.aInd].value
		vToLinksList=[]
		if (vws[getIndex.bInd].value != None):
			vToLinksList.append(vws[getIndex.dInd].value)
		vHasMeta=vws[getIndex.eInd].value
		vws[getIndex.fInd]=5
		vwb.save(vFile)
		vStarRate=vws[getIndex.fInd].value

		row += 1
		getIndex=setIndices(row)
		while ((row <= maxRow) and (vws[getIndex.bInd].value == None) and (vws[getIndex.dInd].value != None)):
			if (vws[getIndex.dInd].value not in vToLinksList):
				vToLinksList.append(vws[getIndex.dInd].value)
			row += 1
			getIndex=setIndices(row)

		ret4StarInfo.retVName=vName
		ret4StarInfo.retVocab=vocab
		ret4StarInfo.retToLinksList=vToLinksList
		ret4StarInfo.retFHasMeta=vHasMeta
		ret4StarInfo.retStarRate=vStarRate
	else:
		ret4StarInfo.retVName=None
		ret4StarInfo.retVocab=None
		ret4StarInfo.retToLinksList=[]
		ret4StarInfo.retFHasMeta=None
		ret4StarInfo.retStarRate=None
	return(ret4StarInfo)
	
def getStarInfo(rws, iStarRate, row):
	class retStarInfo:
		def _init_ (self, retVName, retVocab, retToLinksList, retFHasMeta, retStarRate, retRow):
			self.retVName=""
			self.retVocab=""
			self.retToLinksList=[]
			self.retFHasMeta=False
			self.retStarRate=0
			self.retRow=row
	
	maxRow=rws.max_row
	getIndex=setIndices(row)
	toLinksList=[]

	while ((row <= maxRow) and (rws[getIndex.fInd].value != iStarRate)):
		row += 1
		getIndex=setIndices(row)

	if (rws[getIndex.fInd].value == iStarRate):
		retStarInfo.retVName=rws[getIndex.aInd].value
		if (rws[getIndex.bInd].value == ""):
			retStarInfo.retVocab="No Base URL defined in file"
		else:
			retStarInfo.retVocab=rws[getIndex.bInd].value
		retStarInfo.retFHasMeta=rws[getIndex.eInd].value
		retStarInfo.retStarRate=rws[getIndex.fInd].value

		toLinksList.append(rws[getIndex.dInd].value)
		row += 1
		getIndex=setIndices(row)

		while ((row <= maxRow) and (rws[getIndex.bInd].value == None) and (rws[getIndex.dInd].value != None)):
			if (rws[getIndex.dInd].value not in toLinksList):
				toLinksList.append(rws[getIndex.dInd].value)
			row += 1
			getIndex=setIndices(row)

		retStarInfo.retToLinksList=toLinksList
		retStarInfo.retRow=row
	else:
		retStarInfo.retVName=None
		retStarInfo.retVocab=None
		retStarInfo.retToLinksList=[]
		retStarInfo.retFHasMeta=None
		retStarInfo.retStarRate=None
		retStarInfo.retRow=maxRow+1

	return(retStarInfo)

def writeStarInfo(ws, currentInfo, row):
	getIndex=setIndices(row)

	vToLinksList=currentInfo.retToLinksList

	ws[getIndex.aInd]=currentInfo.retVName
	ws[getIndex.aInd].alignment=Alignment(horizontal='center', wrap_text='True')
	ws[getIndex.bInd]=currentInfo.retVocab
	ws[getIndex.dInd]=currentInfo.retFHasMeta
	ws[getIndex.dInd].alignment=Alignment(horizontal='center', wrap_text='True')
	ws[getIndex.fInd]=currentInfo.retStarRate
	ws[getIndex.fInd].alignment=Alignment(horizontal='center', wrap_text='True')

	currentRow=row
	for item in vToLinksList:
		ws[getIndex.cInd]=item
		currentRow += 1
		getIndex=setIndices(currentRow)
	toLinksRow=currentRow

	return(toLinksRow)

def write5StarsInfo(bFile, bws, rFile, rwb, rws, rws1):
	maxRow=bws.max_row
	currentRow=2
	writeRow=2
	
	while (currentRow <= maxRow):
		getIndex=setIndices(currentRow)
		currentVocab=bws[getIndex.aInd].value
		getFromLinksInfo=collectFromLinksInfo(currentVocab, bFile, bws, currentRow)
		currentRow=getFromLinksInfo.retRow
		getIndex=setIndices(currentRow)
		getUpdateInfo=collect4StarInfo(currentVocab, rFile, rwb, rws)

		writeRow=write5StarSheet(rws1, getUpdateInfo, getFromLinksInfo, writeRow)

	return(writeRow)

def writeOtherStarsInfo(rFile, rws, rws1, writeRow, iStarRate):
	maxRow=rws.max_row
	currentRow=2
	while (currentRow <= maxRow):
		getUpdateInfo=getStarInfo(rws, iStarRate, currentRow)
		currentRow=getUpdateInfo.retRow
		writeRow=writeStarInfo(rws1, getUpdateInfo, writeRow)
	return(writeRow)