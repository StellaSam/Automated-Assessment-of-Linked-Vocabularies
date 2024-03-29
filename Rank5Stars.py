import os
from os.path import isfile, join
import sys

from openpyxl import Workbook
from openpyxl import load_workbook

from VBaseFunctions import *
from UpdateRankedVocabs import *

def chkCurrentWriteRow(row):
	if (row == None):
		row = 2
	return(row)

#fPath="/Users/Stella/Documents/Ph.D. Research/Tool/Tool_Output_SWJ"
#fPath="/Users/Stella/Documents/Ph.D. Research/Tool/Tool_Output_LOV"
#fPath="/Users/Stella/Documents/Ph.D. Research/Tool/Single_Test_Output"
#fPath="/Users/Stella/Documents/Ph.D. Research/Base_URI_Tool/Base_URI_Tool_Output(SWJ)"
fPath="/Users/Stella/Documents/Ph.D. Research/Base_URI_Tool/Base_URI_Tool_Output(LOV)"
#fPath="YOUR OUTPUT FILE LOCATION HERE"
bFileName="vocab_Base.xlsx"
bFile=fPath+"/"+bFileName
bwb=load_workbook(filename=bFile)
bws=bwb["From Links"]

rFileName="Ranked_Vocabs.xlsx"
rFile=fPath+"/"+rFileName
rwb=load_workbook(filename=rFile)
rws=rwb["4-Star Ranked Vocabs"]
rws1=rwb.create_sheet("5-Star Ranked Vocabs", 1)
maxRow=rws.max_row
init5StarSheet(rws1)
currentWriteRow=write5StarsInfo(bFile, bws, rFile, rwb, rws, rws1)
currentWriteRow=chkCurrentWriteRow(currentWriteRow)

currentWriteRow=writeOtherStarsInfo(rFile, rws, rws1, currentWriteRow, "****")
currentWriteRow=chkCurrentWriteRow(currentWriteRow)

currentWriteRow=writeOtherStarsInfo(rFile, rws, rws1, currentWriteRow, "**-*")
currentWriteRow=chkCurrentWriteRow(currentWriteRow)

currentWriteRow=writeOtherStarsInfo(rFile, rws, rws1, currentWriteRow, "***-")
currentWriteRow=chkCurrentWriteRow(currentWriteRow)

currentWriteRow=writeOtherStarsInfo(rFile, rws, rws1, currentWriteRow, "**--")
currentWriteRow=chkCurrentWriteRow(currentWriteRow)	

currentWriteRow=writeOtherStarsInfo(rFile, rws, rws1, currentWriteRow, "*---")
currentWriteRow=chkCurrentWriteRow(currentWriteRow)

rwb.save(rFile)
